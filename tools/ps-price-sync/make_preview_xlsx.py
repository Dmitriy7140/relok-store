# -*- coding: utf-8 -*-
"""
Автономный генератор предпросмотра таблицы (без Google Sheets).

Делает ровно то же, что sync.js, но пишет результат в локальный .xlsx,
который можно скачать и открыть на ПК. Источник цен — ТОЛЬКО официальный
PlayStation Store (store.playstation.com/en-tr).

Запуск:  python tools/ps-price-sync/make_preview_xlsx.py
"""
import os, re, json, time, sqlite3, urllib.request, urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(HERE, '..', '..', 'server', 'data', 'logovo.sqlite')
OUT = os.path.join(HERE, '..', '..', 'Игры_цены_PS_Store_TR.xlsx')

UA = ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
      '(KHTML, like Gecko) Chrome/125.0 Safari/537.36')
THROTTLE = 0.25
RETRIES = 3
MIN_SCORE = 0.55

# ── Издания и синонимы (порт из matcher.js) ────────────────────────────
EDITIONS = {
    'standard': ['standard'],
    'deluxe': ['deluxe', 'digital deluxe'],
    'ultimate': ['ultimate'],
    'gold': ['gold'],
    'premium': ['premium'],
    'complete': ['complete'],
    'goty': ['goty', 'game of the year'],
    'directorscut': ["director's cut", 'directors cut', 'director’s cut'],
}
EDITION_WORDS = sum(EDITIONS.values(), []) + ['edition', 'digital']


def normalize(s):
    s = (s or '').lower()
    s = re.sub(r'[™®©]', ' ', s)
    s = re.sub(r"[:\-–—_,.!?'’\"()\[\]]", ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def detect_edition(text):
    n = normalize(text)
    for cls, words in EDITIONS.items():
        if any(normalize(w) in n for w in words):
            return cls
    return 'standard'


def strip_editions(text):
    n = normalize(text)
    for w in EDITION_WORDS:
        n = n.replace(normalize(w), ' ')
    return re.sub(r'\s+', ' ', n).strip()


def token_sim(a, b):
    sa, sb = set(filter(None, a.split())), set(filter(None, b.split()))
    if not sa or not sb:
        return 0.0
    inter = len(sa & sb)
    return inter / (len(sa) + len(sb) - inter)


def score_pair(game, cand_name):
    gb, cb = strip_editions(game['name']), strip_editions(cand_name)
    s = token_sim(gb, cb)
    if gb and cb and (cb in gb or gb in cb):
        s = max(s, 0.8)
    want = detect_edition(game['edition'] or game['name'])
    got = detect_edition(cand_name)
    s += 0.15 if want == got else -0.2
    return max(0.0, min(1.0, s))


# ── Парсинг цены ───────────────────────────────────────────────────────
def parse_try_number(s):
    if not s:
        return None
    c = re.sub(r'[^\d.,]', '', s)
    if not c:
        return None
    try:
        return float(c.replace('.', '').replace(',', '.'))
    except ValueError:
        return None


def is_free(s):
    return bool(s) and s.strip().lower() in ('free', 'ücretsiz', '0,00 tl')


# ── Извлечение Apollo State (балансировка скобок) ──────────────────────
def extract_apollo(html):
    i = -1
    for marker in ('window.__APOLLO_STATE__', '__APOLLO_STATE__', 'apolloState'):
        p = html.find(marker)
        if p != -1:
            i = html.find('{', p)
            break
    if i == -1:
        return None
    depth = 0
    instr = esc = False
    j = i
    while j < len(html):
        ch = html[j]
        if instr:
            if esc:
                esc = False
            elif ch == '\\':
                esc = True
            elif ch == '"':
                instr = False
        elif ch == '"':
            instr = True
        elif ch == '{':
            depth += 1
        elif ch == '}':
            depth -= 1
            if depth == 0:
                try:
                    return json.loads(html[i:j + 1])
                except json.JSONDecodeError:
                    return None
        j += 1
    return None


def fetch(url):
    last = None
    for attempt in range(1, RETRIES + 1):
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': UA,
                'Accept': 'text/html,application/xhtml+xml',
                'Accept-Language': 'en-TR,en;q=0.9,tr;q=0.8',
            })
            with urllib.request.urlopen(req, timeout=30) as r:
                return r.read().decode('utf-8', 'replace')
        except Exception as e:  # noqa
            last = e
            time.sleep(1.2 * attempt)
    raise last


def search_products(term):
    url = 'https://store.playstation.com/en-tr/search/' + urllib.parse.quote(term)
    state = extract_apollo(fetch(url))
    if not state:
        return []
    out = []
    for v in state.values():
        if not isinstance(v, dict) or v.get('__typename') != 'Product':
            continue
        price = v.get('price')
        if isinstance(price, dict) and '__ref' in price:
            price = state.get(price['__ref'], {})
        ps = None
        if isinstance(price, dict):
            ps = price.get('discountedPrice') or price.get('basePrice')
        out.append({'name': (v.get('name') or '').strip(), 'priceStr': ps})
    return out


def pick_best(game, cands):
    best, best_s = None, 0.0
    for c in cands:
        s = score_pair(game, c['name'])
        if s > best_s:
            best_s, best = s, c
    return (best, best_s) if best and best_s >= MIN_SCORE else (None, best_s)


# ── Основной проход ────────────────────────────────────────────────────
def main():
    con = sqlite3.connect(DB)
    games = [
        {'id': r[0], 'name': (r[1] or '').strip(), 'edition': (r[2] or '').strip(),
         'platform': (r[3] or '').strip(), 'rub': r[4] or 0}
        for r in con.execute(
            "SELECT id,name,edition,platform,price FROM products "
            "WHERE type='game' ORDER BY id")
    ]
    con.close()
    total = len(games)
    print(f'Игр в магазине: {total}', flush=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Цены PS Store TR'
    header = ['ID', 'Название игры', 'Издание', 'Название в PS Store',
              'Цена в турецких лирах (TRY)', 'Цена в рублях (мой магазин)',
              'Статус', 'Обновлено']
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F2436')
        c.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'

    now = time.strftime('%Y-%m-%d %H:%M:%S')
    stats = {}
    for i, g in enumerate(games, 1):
        ps_name, price_str, status = '', '', 'OK'
        try:
            cands = search_products(g['name'])
            if not cands:
                status = 'Требуется проверка'
            else:
                m, sc = pick_best(g, cands)
                if not m:
                    status = 'Требуется проверка'
                else:
                    ps_name = m['name']
                    p = m['priceStr']
                    if is_free(p):
                        price_str, status = 'Ücretsiz', 'Бесплатно'
                    elif not p or parse_try_number(p) is None:
                        status = 'Цена недоступна'
                    else:
                        price_str = p if 'TL' in p.upper() else p
        except Exception as e:  # noqa
            status = 'Ошибка запроса'
            print(f'  [{g["id"]}] {g["name"]}: {e}', flush=True)

        stats[status] = stats.get(status, 0) + 1
        ws.append([g['id'], g['name'], g['edition'], ps_name,
                   price_str, g['rub'], status, now])
        if i % 20 == 0 or i == total:
            print(f'  {i}/{total}…', flush=True)
        time.sleep(THROTTLE)

    # Ширина колонок.
    widths = [6, 34, 18, 40, 26, 24, 20, 20]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = w

    wb.save(OUT)
    print('─── Итоги ───', flush=True)
    for k, v in stats.items():
        print(f'  {k}: {v}', flush=True)
    print('Файл сохранён:', os.path.abspath(OUT), flush=True)


if __name__ == '__main__':
    main()
