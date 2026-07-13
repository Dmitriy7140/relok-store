# -*- coding: utf-8 -*-
"""
Заменяет столбец «Цена в рублях» в готовой таблице предпросмотра
на актуальные цены веб-магазина из Logovo_прайс-лист.xlsx
(лист «Игры», колонка «Цена, ₽»). Сопоставление — по названию,
с уточнением по изданию и платформе.
"""
import os, re
from openpyxl import load_workbook

ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..')
PREVIEW = os.path.join(ROOT, 'Игры_цены_PS_Store_TR.xlsx')
OUT = os.path.join(ROOT, 'Игры_цены_PS_Store_TR_web.xlsx')
PRICELIST = os.path.join(ROOT, 'Logovo_прайс-лист.xlsx')


def norm(s):
    s = (s or '').lower()
    s = re.sub(r'[™®©]', ' ', s)
    s = re.sub(r"[:\-–—_,.!?'’\"()\[\]]", ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


# ── 1. Читаем прайс-лист веб-магазина ──────────────────────────────────
wb_pl = load_workbook(PRICELIST, data_only=True)
ws_pl = wb_pl['Игры']
rows = list(ws_pl.iter_rows(values_only=True))
# Находим строку заголовка (где есть 'Название' и 'Цена, ₽').
hidx = next(i for i, r in enumerate(rows) if r and 'Название' in r)
head = [str(c or '') for c in rows[hidx]]
ci_name = head.index('Название')
ci_plat = head.index('Платформа') if 'Платформа' in head else None
ci_ed = head.index('Издание') if 'Издание' in head else None
ci_price = head.index('Цена, ₽')

# lookup: (name, edition, platform) → цена; плюс запасной по одному имени.
by_full, by_name = {}, {}
for r in rows[hidx + 1:]:
    if not r or not r[ci_name]:
        continue
    name = norm(r[ci_name])
    ed = norm(r[ci_ed]) if ci_ed is not None else ''
    plat = norm(r[ci_plat]) if ci_plat is not None else ''
    price = r[ci_price]
    if price is None:
        continue
    by_full[(name, ed, plat)] = price
    by_name.setdefault(name, price)

print(f'Прайс-лист: {len(by_name)} игр загружено')

# ── 2. Обновляем предпросмотр ──────────────────────────────────────────
wb = load_workbook(PREVIEW)
ws = wb.active
# столбцы: B=Название(2) C=Издание(3) F=Цена в рублях(6)
updated, missed = 0, []
for row in ws.iter_rows(min_row=2):
    name = norm(row[1].value)
    ed = norm(row[2].value)
    # платформу в предпросмотре не храним → пробуем полное совпадение по name+ed,
    # затем по имени.
    price = None
    for key in list(by_full.keys()):
        if key[0] == name and key[1] == ed:
            price = by_full[key]
            break
    if price is None:
        price = by_name.get(name)
    if price is not None:
        row[5].value = price
        updated += 1
    else:
        # Нет в прайс-листе веб-магазина → не показываем устаревшее число.
        row[5].value = 'RUB уточнить'
        missed.append(row[1].value)

wb.save(OUT)
print(f'Обновлено рублёвых цен: {updated}')
if missed:
    print(f'Не найдено в прайс-листе ({len(missed)}): ' + ', '.join(str(m) for m in missed[:40]))
print('Файл сохранён:', os.path.abspath(OUT))
